# -*- coding:utf-8 -*-
# 文件名称：Lyy-Test20201108
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/11/9 11:55

'''
python操作Excel：
    xlrd
    xlwd

    openpyxl：

区别处理文件的长度，xlx，xlxs

'''

import openpyxl
import datetime

Excel_path = r'C:\Users\LyyCc\Desktop\测试是是是.xlsx'
#
# # 创建Excel对象
# wb = openpyxl.Workbook(Excel_path)
#
# # 创建sheet对象
# ws = wb.create_sheet('Sheet1')  # 括号里面是sheet栏的名称，ws是Sheet1的对象
#
# # 保存创建的文件
# wb.save(Excel_path)
#

# 加载Excel
wb = openpyxl.load_workbook(Excel_path)

# 获取表格
ws = wb['Sheet1'] #这里直接读表格，不会报警warning
# ws = wb.get_sheet_by_name('Sheet1')   #这个会在3.9版本报警



data = [{'id': 35859, 'org_id': 263, 'org_code': 'fgjbb', 'creator_id': 36269, 'serial_no': '1', 'range': 'store', 'owner': 'normal', 'desc': '限孕妈礼包领取使用（59.9元孕妈大礼包）', 'c_type': None, 'buz_id': 6, 'plan_action_id': None, 'custom_message': 0, 'message_content': None, 'code_source': 'native', 'name': '418元孕妈礼包优惠卷（满418元减418元）', 'conflict_with_red_envelope': 1, 'standard_buz_id': None, 'custom_template_msg': 'null', 'deputed': 0, 'opposite_code_seq': 2, 'display_in_list': 1, 'state': 'offline', 'global_resource_id': None, 'after_actions': '[]', 'created_at': datetime.datetime(2020, 6, 27, 18, 29, 48), 'updated_at': datetime.datetime(2020, 6, 27, 18, 31, 5), 'plus_only': 0, 'opposite_serial_no': '1', 'influence_journey': 0, 'restriction': '{"scenes": "mall"}', 'coupon_params': '{}', 'product_scope': '{"type": "all", "include_categories": "/"}', 'discount': '{"d_type": 1, "content_value": 418, "condition_value": 418}', 'shop_scope': '{"areas": "", "shops": ""}', 'delivery_methods': '2', 'order_settlement': 'null', 'coupon_type': 'mall', 'icon': None, 'sms_config': '{"enable": false}', 'weixin_template_config': '{"enable": false, "params": [], "use_global": false}', 'communication_task_config': '{"enable": false, "key_points": null, "task_type_id": null, "suggest_words": {"sms": [], "phone": []}, "require_methods": ["phone", "short_msg"]}', 'source_type': 'normal', 'brand_owner_id': None, 'conflict_with_plus_bean': 1, 'extra_options': None}, {'id': 17088, 'org_id': 263, 'org_code': 'fgjbb', 'creator_id': 36271, 'serial_no': 'cp_1552025676161093', 'range': 'all', 'owner': 'standard', 'desc': '测试1元券', 'c_type': None, 'buz_id': 6, 'plan_action_id': None, 'custom_message': 0, 'message_content': None, 'code_source': 'native', 'name': '测试1元券', 'conflict_with_red_envelope': 1, 'standard_buz_id': None, 'custom_template_msg': 'null', 'deputed': 0, 'opposite_code_seq': 1, 'display_in_list': 1, 'state': 'online', 'global_resource_id': None, 'after_actions': '[]', 'created_at': datetime.datetime(2019, 3, 8, 14, 14, 36), 'updated_at': datetime.datetime(2020, 4, 22, 0, 9, 52), 'plus_only': 0, 'opposite_serial_no': 'cp_1552025676161093', 'influence_journey': 0, 'restriction': '{"scenes": "mall"}', 'coupon_params': '{}', 'product_scope': '{"type": "all", "include_categories": "/"}', 'discount': '{"d_type": 1, "content_value": 1, "condition_value": 1}', 'shop_scope': '{"areas": "", "shops": ""}', 'delivery_methods': '2', 'order_settlement': 'null', 'coupon_type': 'mall', 'icon': None, 'sms_config': '{"enable": false}', 'weixin_template_config': '{"enable": false, "params": [], "use_global": false}', 'communication_task_config': '{"enable": false, "key_points": null, "task_type_id": null, "suggest_words": {"sms": [], "phone": []}, "require_methods": ["phone", "short_msg"]}', 'source_type': 'normal', 'brand_owner_id': None, 'conflict_with_plus_bean': 1, 'extra_options': None}]






for i in range(1,len(list(data[0].keys()))+1):
    ws.cell(row=1,column = i,value= list(data[0].keys())[i-1] )

for flag in range(len(data)):
    for i in range(1, len(list(data[flag].values())) + 1):
        ws.cell(row=flag+2, column=i, value=list(data[flag].values())[i - 1])

wb.save(Excel_path)



# 字典


# for i in data[0]:
#     print(i)

# for i in data[0].keys():
#     print(i)

#
# for i in data[0].values():
#     print(i)
#
# for k,v in data[0].items():
#     print(k,v)