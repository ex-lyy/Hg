# -*- coding:utf-8 -*-
# 文件名称：Lyy-去重测试
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/11/2 16:41


'''
从excel里面拿出来第一、二列的数据
打印出两者的交集、存在A不存在B的元素、存在B不存在A的元素
'''

import openpyxl

# 去重文件的绝对路径
Excel_path = r'C:\Users\LyyCc\Desktop\数据去重文件.xlsx'

# 创建一个工作簿
wb = openpyxl.load_workbook(Excel_path)

# 创建一个工作表对象
ws = wb['Sheet1']

# 创建A、B两个列表存储第一、二列的数据  12

column_a = []
column_b = []

# 获取长度最大的列的长度
max_row = ws.max_row

# 循环读取两列的数据，放入对应的表格
for i in range(2, max_row + 2):
    if ws.cell(row=i, column=1).value == None:
        pass
    else:
        column_a.append(ws.cell(row=i, column=1).value)
    if ws.cell(row=i, column=2).value == None:
        pass
    else:
        column_b.append(ws.cell(row=i, column=2).value)

# 输出两列的交集(这里用几何的数据类型最好操作，因为集合有对应函数)
# union_list = set(column_a)&set(column_b)
# print("A和B集的交集为：",'*-'*50)
# for i in union_list:
#     print(i)

A = set(column_a)
B = set(column_b)
print("集合A的长度：",len(A))
print("集合B的长度：",len(B))

# 输出A和B的差集
print("A和B的差集为：", '*-' * 50)

for i in A-B:
    print(i)
