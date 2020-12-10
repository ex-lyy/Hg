# -*- coding:utf-8 -*-
# 文件名称：Hg-lyycc-商户在线情况修正版
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/12/1 17:16

from Mymain import *
import openpyxl
import time
import datetime
from openpyxl.styles import Font, colors, Alignment, PatternFill, Border, Side


# 在业务主库执行查询sql
def execute_sqls(query_sql, result_num, cursor):
    cursor.execute(query_sql)
    if result_num == '1':
        result_data = cursor.fetchone()
        return result_data
    elif result_num == 'all':
        result_data = cursor.fetchall()
        return result_data
    else:
        print("execute_sqls输入的查询参数有误")


# 查询在线的商户信息
def query_online_org():
    query_online_org_sql = "SELECT code,name,profile_id FROM ris_production.global_retailers WHERE profile_id IN ('1','8','9') AND state = 'online' AND service_provider_code='global' AND code NOT IN ('babybear','jinyaolan','happyxybb','loveheart','yyplan','shengyi','beibei','aiwa','jdb','xzmmbb','lyjy','sunnybaby','kkqq','eastbaby','harneybaby','dreamstart','tianyibaby','babyfocus','mamalove','yingyuansu','ybfcbb','qwmykj','zgprettybaby','rsxyzj','rxwayy','lzgoodboy','czqcyy','pxayf','lshibaby','clbabytiandi','wxmyf','fzjiabeiai','fzabd','jzaibaby','sqbaobaole','xzguzi','yqmamalove','glhuangjiayy','pyfujiababy','zyyibaby','bzyqyy','xwmaidibaby','hyqzy','cddiaodiao','gynanwanvwa','gtmamibaby','pzyingyuansu','xpbaobaodangjia','fddamuzhi','ylmamilove','jjdiandian','xmshujie','habaonuoliying','cydoudingzhijia','gzcoolbaby','xcbeierjiazu','pzhkangbaole','mtyunzhiai','lhlvshu','yhaijiabao','mmbabyangel','shbfuwawa','xybuyingfang','czaiyingjiayuan','yzbeibeiyy','czhjyy','jnyft','zbqch','lxdodoisland','cdgeke','lzsimier','hhmbwy','gdxtzgg','zzygts','yckbyy','yycomebaby','aqaxqzy','babyeden','aidingbaby','ssjjw','sevenkea','hzfxbb','lyxbl','pzhappybud','cpbeilejia','masmmbb','scayf','hnqjmy','ssmiergang','aqjbb','xhyyzj','hfyybs','yfxty','qjjbb','qhdbabytown','szxsbb','yymml','yxgayy','nhfbb','wsygbb','ptbbboom','chqzmy','cdbbth','ddjsts','hshjmy','fnqzbb','bbnnbb','kpdcyy','hdayy','axhk','wlbbdj','lyygbb','hgzabb','hmmy');"
    online_org_data_list = execute_sqls(query_online_org_sql, result_num='all', cursor=cursor)
    # 商户数据处理
    online_org_list = []  # 定义org_code的列表
    name_list = []  # 定义商户名称的列表
    profile_id_list = []  # 定义商户类型的列表
    for online_org_data in online_org_data_list:
        online_org_list.append(online_org_data['code'])
        name_list.append(online_org_data['name'])
        if online_org_data['profile_id'] == 1:
            profile_id_list.append('直营')
        elif online_org_data['profile_id'] == 8:
            profile_id_list.append('联营大户')
        elif online_org_data['profile_id'] == 9:
            profile_id_list.append('梧桐')
    return online_org_list, name_list, profile_id_list


# 查询注册信息不正常的商户
def query_register_state(org_code, query_begin_date, cursor):
    register_state_sql = "SELECT COUNT(1) as member_count FROM ris_production.members WHERE org_code='%s' AND created_at>='%s';" % (
        org_code, query_begin_date)
    query_register_num = execute_sqls(register_state_sql, result_num='1', cursor=cursor)
    if query_register_num['member_count'] >= 3:
        return 1
    else:
        return 0


# 查询订单奖励信息不正常的商户
def query_prize_activity_state(org_code, query_begin_date, cursor):
    query_prize_activity_state_sql = "SELECT COUNT(1) as prize_count FROM pomelo_backend_production.prize_activities WHERE org_code='%s' AND created_at>='%s';" % (
        org_code, query_begin_date)
    query_prize_activity_num = execute_sqls(query_prize_activity_state_sql, result_num='1', cursor=cursor)
    if query_prize_activity_num['prize_count'] >= 1:
        return 1
    else:
        return 0


# 查询促销活动信息不正常的商户
def query_promotion_state(org_code, query_begin_date, cursor):
    query_promotion_sql = "SELECT COUNT(1) as promotion_count FROM pomelo_backend_production.promotion_promotions WHERE org_code='%s' AND created_at>='%s';" % (
        org_code, query_begin_date)
    query_promotion_num = execute_sqls(query_promotion_sql, result_num='1', cursor=cursor)
    if query_promotion_num['promotion_count'] >= 1:
        return 1
    else:
        return 0


# 查询互动活动信息不正常的商户
def query_marketing_state(org_code, query_begin_date, cursor):
    query_marketing_sql = "SELECT COUNT(1) as marketing_count FROM pomelo_backend_production.marketing_prize_events WHERE org_code='%s' AND created_at>='%s';" % (
        org_code, query_begin_date)
    query_marketing_num = execute_sqls(query_marketing_sql, result_num='1', cursor=cursor)
    if query_marketing_num['marketing_count'] >= 1:
        return 1
    else:
        return 0


# 创建Excel表格
def create_excel(Excel_path):
    # 先创建一个表格
    wb_create = openpyxl.Workbook(Excel_path)
    wb_create.create_sheet('业务侧异常商户信息')
    wb_create.create_sheet('业务侧异常详情')
    wb_create.create_sheet('在线运营商户信息')
    wb_create.save(Excel_path)
    return 0


# 写入标准标题信息
def write_excel_demo(Excel_path):
    wb_demo = openpyxl.load_workbook(Excel_path)
    ws_demo_one = wb_demo['在线运营商户信息']
    ws_demo_two = wb_demo['业务侧异常商户信息']
    ws_demo_three = wb_demo['业务侧异常详情']
    # 写表格的头
    ws_demo_one.cell(row=1, column=1, value='商户名称')
    ws_demo_one.cell(row=1, column=2, value='商户编号')
    ws_demo_one.cell(row=1, column=3, value='商户类型')
    ws_demo_two.cell(row=1, column=1, value='商户名称')
    ws_demo_two.cell(row=1, column=2, value='商户编号')
    ws_demo_two.cell(row=1, column=3, value='商户类型')
    ws_demo_two.cell(row=1, column=4, value='发现异常时间')
    ws_demo_three.cell(row=1, column=1, value='会员升级异常商户（近3月升级会员小于3人）')
    ws_demo_three.cell(row=1, column=3, value='订单奖励活动（近3月无活动）')
    ws_demo_three.cell(row=1, column=5, value='促销（近3月无活动）')
    ws_demo_three.cell(row=1, column=7, value='互动游戏（近3月无活动）')
    for column in 'ABCDEFGH':
        ws_demo_one.column_dimensions[column].width = 20
        ws_demo_two.column_dimensions[column].width = 20
        ws_demo_three.column_dimensions[column].width = 25
    wb_demo.save(Excel_path)
    return 0

def write_normal_data(Excel_path, online_org_list, name_list, profile_id_list):
    wb = openpyxl.load_workbook(Excel_path)
    ws = wb['在线运营商户信息']
    for i in range(0, len(online_org_list)):
        ws.cell(row=i + 2, column=1, value=online_org_list[i])
        ws.cell(row=i + 2, column=2, value=name_list[online_org_list.index(online_org_list[i])])
        ws.cell(row=i + 2, column=3, value=profile_id_list[online_org_list.index(online_org_list[i])])
    wb.save(Excel_path)
    return 0


def write_unnormal_data(Excel_path, unnormal_org_list, online_org_list, name_list, profile_id_list):
    wb = openpyxl.load_workbook(Excel_path)
    ws = wb['业务侧异常商户信息']
    # 写当月第一天的日期
    now = datetime.datetime.now()
    this_month_start = datetime.datetime(now.year, now.month, 1).strftime("%Y年%m月")
    for i in range(0, len(unnormal_org_list)):
        ws.cell(row=i + 2, column=1, value=unnormal_org_list[i])
        ws.cell(row=i + 2, column=2, value=name_list[online_org_list.index(unnormal_org_list[i])])
        ws.cell(row=i + 2, column=3, value=profile_id_list[online_org_list.index(unnormal_org_list[i])])
        ws.cell(row=i + 2, column=4, value=this_month_start)
    wb.save(Excel_path)
    return 0


def write_unnormal_activity_data(Excel_path, unnormal_member_org_list, unnormal_prize_org_list,
                                 unnormal_promotion_org_list, unnormal_marketing_org_list, online_org_list, name_list):
    wb = openpyxl.load_workbook(Excel_path)
    ws = wb['业务侧异常详情']
    for i in range(0, len(unnormal_member_org_list)):
        ws.cell(row=i + 2, column=1, value=unnormal_member_org_list[i])
        ws.cell(row=i + 2, column=2, value=name_list[online_org_list.index(unnormal_member_org_list[i])])
    for i in range(0, len(unnormal_prize_org_list)):
        ws.cell(row=i + 2, column=3, value=unnormal_prize_org_list[i])
        ws.cell(row=i + 2, column=4, value=name_list[online_org_list.index(unnormal_prize_org_list[i])])
    for i in range(0, len(unnormal_promotion_org_list)):
        ws.cell(row=i + 2, column=5, value=unnormal_promotion_org_list[i])
        ws.cell(row=i + 2, column=6, value=name_list[online_org_list.index(unnormal_promotion_org_list[i])])
    for i in range(0, len(unnormal_marketing_org_list)):
        ws.cell(row=i + 2, column=7, value=unnormal_marketing_org_list[i])
        ws.cell(row=i + 2, column=8, value=name_list[online_org_list.index(unnormal_marketing_org_list[i])])
    wb.save(Excel_path)
    return 0


start_time = time.time()
# 定义查询开始时间
query_begin_date = '2020-09-01'

if __name__ == '__main__':
    # 会员不正常的商户：
    unnormal_member_org_list = []
    # 订单奖励数量不正常的商户：
    unnormal_prize_org_list = []
    # 促销活动奖励数量不正常的商户：
    unnormal_promotion_org_list = []
    # 抽奖活动数量不正常的商户：
    unnormal_marketing_org_list = []
    # 认定为不正常的商户
    unnormal_org_list = []
    # 开启数据库链接：
    server, dbconfig, cursor = connect_master_copy_DB('ris_production')
    # 查询在线运营商户
    online_org_list, name_list, profile_id_list = online_org_data = query_online_org()
    for org_code in online_org_list:
        print("正在处理%s商户的数据！" % org_code)

        # 查询注册会员数量是否正常
        register_unnormal_org_state = query_register_state(org_code,
                                                            query_begin_date, cursor)
        prize_activity_state = query_prize_activity_state(org_code, query_begin_date, cursor)
        promotion_activity_state = query_promotion_state(org_code, query_begin_date, cursor)
        marketing_activity_state = query_prize_activity_state(org_code, query_begin_date, cursor)
        if register_unnormal_org_state == 0:
            unnormal_member_org_list.append(org_code)
        if prize_activity_state == 0:
            unnormal_prize_org_list.append(org_code)
        if promotion_activity_state == 0:
            unnormal_promotion_org_list.append(org_code)
        if marketing_activity_state == 0:
            unnormal_marketing_org_list.append(org_code)
        if register_unnormal_org_state == prize_activity_state == promotion_activity_state == marketing_activity_state == 0:
            unnormal_org_list.append(org_code)
    # 关闭数据库链接
    close_sshserver(server, dbconfig, cursor)

    # 创建表格路径
    Excel_path = r'C:\Users\LyyCc\Desktop\业务侧异常商户情况' + query_begin_date + r'.xlsx'
    create_excel(Excel_path)
    write_excel_demo(Excel_path)
    write_normal_data(Excel_path, online_org_list, name_list, profile_id_list)
    write_unnormal_data(Excel_path, unnormal_org_list, online_org_list, name_list, profile_id_list)
    write_unnormal_activity_data(Excel_path, unnormal_member_org_list, unnormal_prize_org_list,
                                 unnormal_promotion_org_list, unnormal_marketing_org_list, online_org_list, name_list)
    end_time = time.time()
    print("本次操作用时：", end_time - start_time)
