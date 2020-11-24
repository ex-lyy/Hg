# -*- coding:utf-8 -*-
# 文件名称：Hg-lyycc-大转盘标准化检查
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/11/24 10:09

from Mymain import *
import openpyxl
from openpyxl.styles import Font, colors, Alignment,PatternFill,Border, Side

def create_excel(result_excel_path):
    # 生成指定路径的Excel文件(此步骤只能进行生成工作，本次的表格对象是只读模式)
    print(result_excel_path)
    wb = openpyxl.Workbook(result_excel_path)
    ws = wb.create_sheet('抽奖活动标准化检查结果')
    wb.save(result_excel_path)
    return 0

def load_excel(result_excel_path):
    print(result_excel_path)
    wb = openpyxl.load_workbook(result_excel_path)
    ws = wb['抽奖活动标准化检查结果']
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10)
    ws.merge_cells(start_row=19, start_column=1, end_row=18, end_column=10)
    ws.merge_cells(start_row=30, start_column=1, end_row=32, end_column=10)
    cell_list = ['1','5','19','30']
    ws['A1'] = '活动基本信息'
    ws['A5'] = '奖励概率及数量信息'
    ws['A19'] = '配置的奖励信息'
    ws['A32'] = '奖励等级信息'

    ws = make_cell_style(cell_list,ws)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A19'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A32'].alignment = Alignment(horizontal='center', vertical='center')
    # 写模板数据-活动基本信息
    ws.cell(row= 2 ,column= 1,value='ID')
    ws.cell(row= 2 ,column= 2,value='商户code')
    ws.cell(row= 2 ,column= 3,value='活动code')
    ws.cell(row= 2 ,column= 4,value='活动名称')
    ws.cell(row= 2 ,column= 5,value='活动开始时间')
    ws.cell(row= 2 ,column= 6,value='活动结束时间')
    # 写模板数据-中奖概率及数量
    ws.cell(row= 6 ,column= 1,value='奖励序号')
    ws.cell(row= 6 ,column= 2,value='奖励ID')
    ws.cell(row= 6 ,column= 3,value='概率（百分比）')
    ws.cell(row= 6 ,column= 4,value='数量')
    ws.cell(row= 6 ,column= 5,value='奖项名称')
    ws.cell(row= 6 ,column= 6,value='是否是备份奖励')
    # 写模板数据-配置的奖励信息
    ws.cell(row= 20 ,column= 1,value='奖励类型')
    ws.cell(row= 20 ,column= 2,value='奖励名称')
    ws.cell(row= 20 ,column= 3,value='中奖描述')
    ws.cell(row= 20 ,column= 4,value='奖励编号')
    ws.cell(row= 20 ,column= 5,value='固定天数')
    ws.cell(row= 20 ,column= 6,value='奖励开始时间')
    ws.cell(row= 20 ,column= 7,value='结束时间')
    # 写模板数据-写奖励等级信息
    ws.cell(row= 33 ,column= 1,value='type')
    ws.cell(row= 33 ,column= 2,value='活动id')
    ws.cell(row= 33 ,column= 3,value='等级id')
    ws.cell(row= 33 ,column= 4,value='奖励')


    wb.save(result_excel_path)
    return 0

# 定义顶层模板样式
def make_cell_style(cell_list,ws):
    cell_row_list = ['A','B','C','D','E','F','G','H','I','J']
    cell_font = Font(u'宋体',size = 11,bold=True,italic=False,strike=False)
    cell_fill = PatternFill("solid", fgColor="ffa64d")
    cell_border = Border(top=Side(border_style='thin', color=colors.BLACK),
                    bottom=Side(border_style='thin', color=colors.BLACK),
                    left=Side(border_style='thin', color=colors.BLACK),
                    right=Side(border_style='thin', color=colors.BLACK))
    for cell in cell_list:
        for cell_row in cell_row_list:
            ws[cell_row+cell].font = cell_font
            ws[cell_row+cell].fill = cell_fill
            ws[cell_row+cell].border = cell_border
            ws.column_dimensions[cell_row].width = 20
    return ws


def write_data(wb,ws):
    pass


org_code = 'dyxinya'
activity_code = 'pe_1589620826768826'

excel_name_time = get_time_str()
# 存放结果的文件的绝对路径
result_excel_path = r'C:\Users\LyyCc\Desktop\抽奖标准化-%s.xlsx' %excel_name_time
# 有对应模板的excel文件执行load_excel,如果没有执行create_excel+load_excel
print(result_excel_path)
create_excel(result_excel_path)
load_excel(result_excel_path)

def query_base_info():
    server, dbconfig, cursor = connect_master_copy_DB('pomelo_backend_production')

    base_info_query_sql = "SELECT id,org_code,code,name,begin_time,end_time FROM pomelo_backend_production.marketing_prize_events WHERE org_code='%s' AND code='%s';" % (
    org_code, activity_code)
    cursor.execute(base_info_query_sql)
    base_info = cursor.fetchone()
    print(base_info)
    close_sshserver(server, dbconfig, cursor)
    return base_info