# -*- coding:utf-8 -*-
# 文件名称：Hg-lyycc-大转盘标准化检查
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/11/24 10:09

from Mymain import *
import openpyxl
from openpyxl.styles import Font, colors, Alignment,PatternFill,Border, Side
import re

def create_excel(result_excel_path):
    # 生成指定路径的Excel文件(此步骤只能进行生成工作，本次的表格对象是只读模式)
    wb = openpyxl.Workbook(result_excel_path)
    ws1 = wb.create_sheet('抽奖活动标准化检查结果')
    ws2 = wb.create_sheet('抽奖活动标准化检查描述')
    wb.save(result_excel_path)
    return 0

def write_medoel_excel(result_excel_path):
    wb = openpyxl.load_workbook(result_excel_path)
    ws = wb['抽奖活动标准化检查结果']
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=10)
    ws.merge_cells(start_row=19, start_column=1, end_row=19, end_column=10)
    ws.merge_cells(start_row=32, start_column=1, end_row=32, end_column=10)
    cell_list = ['1','5','19','32']
    ws['A1'] = '活动基本信息'
    ws['A5'] = '奖励概率及数量信息'
    ws['A19'] = '配置的奖励信息'
    ws['A32'] = '奖励等级信息'

    ws = make_cell_style(cell_list,ws)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A19'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A32'].alignment = Alignment(horizontal='center', vertical='center')
    # 写模板数据-活动基本信息
    ws.cell(row= 2 ,column= 1,value='ID')
    ws.cell(row= 2 ,column= 2,value='商户code')
    ws.cell(row= 2 ,column= 3,value='活动code')
    ws.cell(row= 2 ,column= 4,value='活动名称')
    ws.cell(row= 2 ,column= 5,value='活动开始时间')
    ws.cell(row= 2 ,column= 6,value='活动结束时间')
    # 写模板数据-中奖概率及数量
    ws.cell(row= 6 ,column= 1,value='奖励序号')
    ws.cell(row= 6 ,column= 2,value='奖励ID')
    ws.cell(row= 6 ,column= 3,value='概率（百分比）')
    ws.cell(row= 6 ,column= 4,value='数量')
    ws.cell(row= 6 ,column= 5,value='奖项名称')
    ws.cell(row= 6 ,column= 6,value='是否是备份奖励')
    # 写模板数据-配置的奖励信息
    ws.cell(row= 20 ,column= 1,value='奖励类型')
    ws.cell(row= 20 ,column= 2,value='奖励名称')
    ws.cell(row= 20 ,column= 3,value='中奖描述')
    ws.cell(row= 20 ,column= 4,value='奖励编号')
    ws.cell(row= 20 ,column= 5,value='固定天数')
    ws.cell(row= 20 ,column= 6,value='奖励开始时间')
    ws.cell(row= 20 ,column= 7,value='结束时间')
    ws.cell(row= 20 ,column= 8,value='最大红包金额')
    ws.cell(row= 20 ,column= 9,value='最小红包金额')
    # 写模板数据-写奖励等级信息
    ws.cell(row= 33 ,column= 1,value='type')
    ws.cell(row= 33 ,column= 2,value='活动id')
    ws.cell(row= 33 ,column= 3,value='等级id')
    ws.cell(row= 33 ,column= 4,value='奖励')
    wb.save(result_excel_path)
    return 0

# 定义顶层模板样式
def make_cell_style(cell_list,ws):
    cell_row_list = ['A','B','C','D','E','F','G','H','I','J']
    cell_font = Font(u'宋体',size = 11,bold=True,italic=False,strike=False)
    cell_fill = PatternFill("solid", fgColor="ffa64d")
    cell_border = Border(top=Side(border_style='thin', color=colors.BLACK),
                    bottom=Side(border_style='thin', color=colors.BLACK),
                    left=Side(border_style='thin', color=colors.BLACK),
                    right=Side(border_style='thin', color=colors.BLACK))
    for cell in cell_list:
        for cell_row in cell_row_list:
            ws[cell_row+cell].font = cell_font
            ws[cell_row+cell].fill = cell_fill
            ws[cell_row+cell].border = cell_border
            ws.column_dimensions[cell_row].width = 27.5
    return ws

def query_base_info(org_code, activity_code):
    server, dbconfig, cursor = connect_master_copy_DB('pomelo_backend_production')
    base_info_query_sql = "SELECT id,org_code,code,name,begin_time,end_time FROM pomelo_backend_production.marketing_prize_events WHERE org_code='%s' AND code='%s';" % (
    org_code, activity_code)
    cursor.execute(base_info_query_sql)
    base_info = cursor.fetchone()
    close_sshserver(server, dbconfig, cursor)
    return base_info,base_info['id'],base_info['name'],base_info['begin_time'],base_info['end_time']


def query_count_info(activity_id):
    server, dbconfig, cursor = connect_master_copy_DB('pomelo_backend_production')
    base_info_query_sql = "SELECT seq,IFNULL(id, '总计') id,CONCAT(ROUND(SUM(probability),2),'%%') AS 概率,SUM(num) AS 数量,prize_name AS 奖品,backup As 备份奖励 FROM pomelo_backend_production.marketing_random_prizes WHERE prize_event_id = %s GROUP BY id WITH ROLLUP;" %activity_id
    cursor.execute(base_info_query_sql)
    base_info = cursor.fetchall()
    close_sshserver(server, dbconfig, cursor)
    return base_info

def query_serial_name(org_code,serial_no):
    server, dbconfig, cursor = connect_master_copy_DB('pomelo_backend_production')
    base_info_query_sql = "SELECT name FROM pomelo_backend_production.promotion_coupon_definitions WHERE org_code='%s' AND serial_no='%s';" %(org_code,serial_no)
    print(base_info_query_sql)
    cursor.execute(base_info_query_sql)
    base_info_data = cursor.fetchone()

    close_sshserver(server, dbconfig, cursor)
    return  base_info_data['name']


def query_prize_info(activity_id):
    server, dbconfig, cursor = connect_master_copy_DB('pomelo_backend_production')
    base_info_query_sql = "SELECT * FROM pomelo_backend_production.marketing_random_prizes WHERE prize_event_id=%s;" %activity_id
    cursor.execute(base_info_query_sql)
    base_info_data = cursor.fetchall()
    close_sshserver(server, dbconfig, cursor)
    write_lines_data_list = []
    for base_info in base_info_data:
        write_lines_data_dict = {}
        prize_type = expiration_day = begin_date = end_date = serial_no = prize_name = max_price = min_price =prize_item_desc =  '-'
        actions = eval(base_info['actions'])
        if len(actions) == 0:
            pass
        elif len(actions) == 1 :
            actions = actions[0]
            if actions['type']== 'coupon':
                prize_type = '优惠券'
                if actions['params']['expiration_date_type'] == 'by_day':
                    expiration_day = actions['params']['expiration_day']
                    begin_date =''
                    end_date = ''
                    serial_no = actions['params']['serial_no']
                    prize_name = query_serial_name(org_code,serial_no)
                    try:
                        prize_item_desc = actions['desc']
                    except KeyError:
                        prize_item_desc = '-'
                elif actions['params']['expiration_date_type'] == 'by_date':
                    expiration_day = ''
                    begin_date = actions['params']['begin_date']
                    end_date = actions['params']['end_date']
                    serial_no = actions['params']['serial_no']
                    prize_name = query_serial_name(org_code, serial_no)
                    try:
                        prize_item_desc = actions['desc']
                    except KeyError:
                        prize_item_desc = '-'
            elif actions['type']== 'ticket':
                prize_type ='兑换券'
                if actions['params']['expiration_date_type'] == 'by_day':
                    expiration_day = actions['params']['expiration_day']
                    begin_date =''
                    end_date = ''
                    serial_no = actions['params']['activity_id']
                    prize_name = actions['prize_name']
                    try:
                        prize_item_desc = actions['desc']
                    except KeyError:
                        prize_item_desc = '-'
                elif actions['params']['expiration_date_type'] == 'by_date':
                    expiration_day = ''
                    begin_date = actions['params']['begin_date']
                    end_date = actions['params']['end_date']
                    serial_no = actions['params']['activity_id']
                    prize_name = actions['prize_name']
                    try:
                        prize_item_desc = actions['desc']
                    except KeyError:
                        prize_item_desc = '-'
            elif actions['type']== 'package':
                prize_type ='画像礼包'
                serial_no=actions['params']['code']
                prize_name = actions['prize_name']
                try:
                    prize_item_desc = actions['desc']
                except KeyError:
                    prize_item_desc = '-'
            elif actions['type']== 'checkin_card':
                    expiration_date =actions['params']['expiration_date']
                    card_days = actions['params']['card_days']
                    try:
                        prize_item_desc = actions['desc']
                    except KeyError:
                        prize_item_desc = '-'
            elif actions['type']== 'coupon_bag':
                prize_type ='礼券包'
                if actions['params']['expiration_date_type'] == 'by_day':
                    expiration_day = actions['params']['expiration_day']
                    begin_date = ''
                    end_date = ''
                    serial_no = actions['params']['code']
                    prize_name = actions['prize_name']
                    try:
                        prize_item_desc = actions['desc']
                    except KeyError:
                        prize_item_desc = '-'
                elif actions['params']['expiration_date_type'] == 'by_date':
                    expiration_day = ''
                    begin_date = actions['params']['begin_date']
                    end_date = actions['params']['end_date']
                    serial_no = actions['params']['code']
                    prize_name = actions['prize_name']
                    try:
                        prize_item_desc = actions['desc']
                    except KeyError:
                        prize_item_desc = '-'
            elif actions['type']== 'intelligent_package':
                prize_type ='智能礼包'
                serial_no=actions['params']['code']
                try:
                    prize_item_desc = actions['desc']
                except KeyError:
                    prize_item_desc = '-'
            elif actions['type']== 'red_pack':
                prize_type ='微信红包'
                max_price =  actions['params']['max_price']
                min_price =  actions['params']['min_price']
                prize_name = actions['prize_name']
                try:
                    prize_item_desc = actions['desc']
                except KeyError:
                    prize_item_desc = '-'
            elif actions['type']== 'red_envelope':
                prize_type = '商城红包'
                max_price = actions['params']['max_price']
                min_price = actions['params']['min_price']
                prize_name = actions['prize_name']
                try:
                    prize_item_desc = actions['desc']
                except KeyError:
                    prize_item_desc = '-'
            else:
                pass
        else:
            print("有奖项配置的奖励超过一个，不适用本次检测！")
        write_lines_data_dict['prize_type'] = prize_type
        write_lines_data_dict['prize_name'] = prize_name
        write_lines_data_dict['describe'] = prize_item_desc
        write_lines_data_dict['serial_no'] = serial_no
        write_lines_data_dict['expiration_day'] = expiration_day
        write_lines_data_dict['begin_date'] = begin_date
        write_lines_data_dict['end_date'] = end_date
        write_lines_data_dict['max_price'] = max_price
        write_lines_data_dict['min_price'] = min_price
        write_lines_data_list.append(write_lines_data_dict)
    return write_lines_data_list

def query_garde_level(activity_id):
    server, dbconfig, cursor = connect_master_copy_DB('pomelo_backend_production')
    base_info_query_sql = "SELECT c.type,c.activity_id,g.id AS grade_id,g.prizes FROM pomelo_backend_production.business_support_conditions c LEFT JOIN pomelo_backend_production.business_support_condition_grades g ON c.id = g.condition_id WHERE c.activity_id='%s'"%activity_id
    cursor.execute(base_info_query_sql)
    base_info_data = cursor.fetchall()
    close_sshserver(server, dbconfig, cursor)
    return base_info_data

def write_data(result_excel_path,data,write_column_row):
    wb = openpyxl.load_workbook(result_excel_path)
    ws = wb['抽奖活动标准化检查结果']
    if isinstance(data,dict):
        write_column_index = 1
        write_column_row_inside = write_column_row
        for cell_data in data.values():
            ws.cell(row=write_column_row_inside,column=write_column_index,value=cell_data)
            write_column_index += 1
    elif isinstance(data,list):
        write_column_row_inside = write_column_row
        for cell_data_dict in data:
            write_column_index = 1
            for cell_data in cell_data_dict.values():
                ws.cell(row=write_column_row_inside, column=write_column_index, value=cell_data)
                write_column_index += 1
            write_column_row_inside += 1
    else:
        print("写入奖励概率及数量信息---有问题")
    wb.save(result_excel_path)
    return 0

def write_describe(org_code,activity_code,activity_name,activity_begin_date,activity_end_date,result_excel_path):
    check_url = r"https://%s.w.joowing.com/org/%s/prize_events/%s/play?c_type=62&c_code=%s&"%(org_code,org_code,activity_code,activity_code)
    describe = f'''1.测试活动展示界面（附一张活动界面的图，附一张NB奖励抽奖记录的图）
  1.1. 审核状态下给自己发多次抽奖机会，测试转盘图片、指针是否正确 已检查，无误
  1.2. 测试活动下发的奖励是否和获奖等级对应 已检查，无误
  1.3. 测试各奖项的概率大致是否一样，不应出现概率最小的奖项前几次就被抽中 已检查，无误
  1.4. 测试活动限制次数是否正常 已检查，无误
2.活动主题、活动时间
    {activity_name}[{activity_code}]
    {activity_begin_date} ~ {activity_end_date}
3.活动配置好之后中奖概率和份数 OK
4.各个奖励是否正确配置 OK
5.活动参与方式（什么机会可抽什么奖励）
  【需要手动填写】
6.分享参数(自己进到活动界面分享一次确认分享信息正确) OK
7.活动参与方式，检查sql，,activity_id是步骤2里面查询出来的id
8.活动审核链接：
{check_url}'''
    wb = openpyxl.load_workbook(result_excel_path)
    ws = wb['抽奖活动标准化检查描述']
    ws.cell(row=1,column=1,value=describe)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.column_dimensions['A'].width = 100
    ws.row_dimensions[1].hegiht = 100
    wb.save(result_excel_path)
    return  0

if __name__ == '__main__':
    # 手动输入活动数据：
    org_code = str(input("请输入商户code："))
    activity_code = str(input("请输入活动code："))


    excel_name_time = get_time_str()
    # 存放结果的文件的绝对路径
    result_excel_path = r'C:\Users\LyyCc\Desktop\抽奖标准化-%s.xlsx' %excel_name_time
    # 有对应模板的excel文件执行load_excel,如果没有执行create_excel+load_excel
    print("当前工作文件路径为：",result_excel_path)    # 输出当前文件路径
    # 创建文件并写入标准模板
    create_excel(result_excel_path)
    write_medoel_excel(result_excel_path)

    # 写入活动基本信息
    write_column_row = 3    #3是写入表格的行位置，本次做默认
    base_info_data,activity_id,activity_name,activity_begin_date,activity_end_date= query_base_info(org_code, activity_code)
    write_data(result_excel_path,base_info_data,write_column_row)

    # 写入奖励概率及数量信息
    write_column_row = 7    #7是写入表格的行位置，本次做默认
    activity_count_info = query_count_info(activity_id)
    write_data(result_excel_path,activity_count_info,write_column_row)

    # 写入配置的奖励信息（目前只支持一个奖项一个奖励）
    write_column_row = 21    #7是写入表格的行位置，本次做默认
    activity_prize_info = query_prize_info(activity_id)
    write_data(result_excel_path,activity_prize_info,write_column_row)

    # 写入配置的活动等级信息（目前只支持一个奖项一个奖励）
    write_column_row = 34    #7是写入表格的行位置，本次做默认
    query_garde_level_info = query_garde_level(activity_id)
    write_data(result_excel_path,query_garde_level_info,write_column_row)

    # 写入Sheet2栏活动描述
    write_describe(org_code,activity_code,activity_name,activity_begin_date,activity_end_date,result_excel_path)

    print("本次操作完成，请认真核对数据！")