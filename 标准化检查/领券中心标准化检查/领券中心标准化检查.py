# -*- coding:utf-8 -*-
# 文件名称：Hg-lyycc-领券中心标准化检查
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/12/7 10:41

import openpyxl
from Mymain import *
from openpyxl.styles import Font, colors, Alignment,PatternFill,Border, Side


def create_excel(result_excel_path):
    # 生成指定路径的Excel文件(此步骤只能进行生成工作，本次的表格对象是只读模式)
    wb = openpyxl.Workbook(result_excel_path)
    ws1 = wb.create_sheet('领券中心标准化检查结果')
    ws2 = wb.create_sheet('领券中心标准化检查描述')
    wb.save(result_excel_path)
    return 0

# 定义顶层模板样式
def make_cell_style(ws):
    cell_row_list = ['A','B','C','D','E','F','G','H','I','J','K','L']
    cell_list = ['1','5']
    cell_font = Font(u'宋体',size = 11,bold=True,italic=False,strike=False)
    cell_fill = PatternFill("solid", fgColor="ffa64d")
    cell_border = Border(top=Side(border_style='thin', color=colors.BLACK),
                    bottom=Side(border_style='thin', color=colors.BLACK),
                    left=Side(border_style='thin', color=colors.BLACK),
                    right=Side(border_style='thin', color=colors.BLACK))
    for cell in cell_list:
        for cell_row in cell_row_list:
            ws[cell_row+cell].font = cell_font
            ws[cell_row+cell].fill = cell_fill
            ws[cell_row+cell].border = cell_border
            ws.column_dimensions[cell_row].width = 25
    cell_font_title = Font(u'宋体', size=14, bold=True, italic=False, strike=False)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = cell_font_title
    ws['A5'].font = cell_font_title

    return 0

# 写Excel分区标题
def write_medoel_excel(result_excel_path):
    wb = openpyxl.load_workbook(result_excel_path)
    ws = wb['领券中心标准化检查结果']
    first_title = ['活动ID','活动名称','活动编号','活动描述','活动开始时间','活动结束时间','是否周期性','周期性活动时间','可领取数量','是否新客专享','是否可从商品详情界面领券','是否支持分享']
    second_title = ['限制类型','限制次数','奖励类型','奖励编号','奖励名称','固定天数','开始日期','结束日期','限制领取类型','限制领取条码']
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=12)
    ws['A1'] = '活动基本信息'
    ws['A5'] = '配置的奖励信息'
    for i in range(len(first_title)):
        ws.cell(row = 2,column= 1+ i,value=first_title[i])
    for i in range(len(second_title)):
        ws.cell(row = 6,column= 1+ i,value=second_title[i])
    make_cell_style(ws)
    wb.save(result_excel_path)
    return 0

# 查询活动基本信息
def query_avtivity_base_info(org_code,activity_code):
    query_avtivity_base_info_sql= "SELECT id,name,code,description,begin_date,end_date,cycle_type,active_dates,limit_config,new_member_able,visibility,shareable FROM pomelo_backend_production.rush_to_get_coupon_activities WHERE org_code='%s' AND code='%s';"%(org_code,activity_code)
    server, dbconfig, cursor = connect_master_copy_DB('pomelo_backend_production')
    cursor.execute(query_avtivity_base_info_sql)
    avtivity_base_info_data = cursor.fetchone()
    close_sshserver(server, dbconfig, cursor)
    if avtivity_base_info_data['cycle_type']== 'none':
        avtivity_base_info_data['limit_config'] =  eval(avtivity_base_info_data['limit_config'])['total']
        return avtivity_base_info_data
    else:
        print("周期性领券中心，暂时未处理")

# 查询配置的奖励信息
def query_avtivity_prize_info(org_code, activity_id):
    query_avtivity_prize_info_sql = "select reward_limit,max_count,prize_type,prize_info,member_restriction_count,member_restriction,restriction_product_skus from pomelo_backend_production.rush_to_get_coupon_prize_items where org_code='%s' and activity_id= '%s' order by priority desc;" % (
    org_code, activity_id)
    server, dbconfig, cursor = connect_master_copy_DB('pomelo_backend_production')
    cursor.execute(query_avtivity_prize_info_sql)
    activity_prize_info_data = cursor.fetchall()
    close_sshserver(server, dbconfig, cursor)
    activity_prize_item_list = []
    for activity_prize_item in activity_prize_info_data:
        limit_type = limit_count = prize_type = prize_serial_no = prize_serial_name = expiration_date = prize_begin_date = prize_end_date = prize_get_limit_type = prize_get_limit_skus = '-'
        activity_prize_item_dict = {}
        # 划分限制类型
        if activity_prize_item['reward_limit'] == 'total':
            limit_type = '限制总量'
        elif activity_prize_item['reward_limit'] == 'none':
            limit_type = '不限制'
        elif activity_prize_item['reward_limit'] == 'day':
            limit_type = '每日限量'
        limit_count = activity_prize_item['max_count']
        # 划分奖励类型
        if activity_prize_item['prize_type'] == 'coupon':
            prize_type = '优惠券'
            activity_prize_item_info = json.loads(activity_prize_item['prize_info'])
            if activity_prize_item_info['params']['expiration_date_type'] == 'by_day':
                expiration_date = activity_prize_item_info['params']['expiration_date']
            elif activity_prize_item_info['params']['expiration_date_type'] == 'by_date':
                prize_begin_date = activity_prize_item_info['params']['begin_date']
                prize_end_date = activity_prize_item_info['params']['end_date']
            prize_serial_no = activity_prize_item_info['params']['serial_no']
            prize_serial_name = activity_prize_item_info['prize_name']
        elif activity_prize_item['prize_type'] == 'ticket':
            prize_type = '兑换券'
            activity_prize_item_info = json.loads(activity_prize_item['prize_info'])
            if activity_prize_item_info['params']['expiration_date_type'] == 'by_day':
                expiration_date = activity_prize_item_info['params']['expiration_date']
            elif activity_prize_item_info['params']['expiration_date_type'] == 'by_date':
                prize_begin_date = activity_prize_item_info['params']['begin_date']
                prize_end_date = activity_prize_item_info['params']['end_date']
            prize_serial_no = activity_prize_item_info['id']
            prize_serial_name = activity_prize_item_info['params']['name']
        elif activity_prize_item['prize_type'] == 'coupon_bag':
            prize_type = '礼券包'
            activity_prize_item_info = json.loads(activity_prize_item['prize_info'])
            if activity_prize_item_info['params']['expiration_date_type'] == 'by_day':
                expiration_date = activity_prize_item_info['params']['expiration_date']
            elif activity_prize_item_info['params']['expiration_date_type'] == 'by_date':
                prize_begin_date = activity_prize_item_info['params']['begin_date']
                prize_end_date = activity_prize_item_info['params']['end_date']
            prize_serial_no = activity_prize_item_info['params']['code']
            prize_serial_name = activity_prize_item_info['prize_name']
        # 判断券领取限制
        if activity_prize_item['member_restriction'] == 'all':
            prize_get_limit_type = '不限制'
        elif activity_prize_item['member_restriction'] == 'not_buy_product':
            prize_get_limit_type = '限制未购条码'
            prize_get_limit_skus = activity_prize_item['restriction_product_skus']
        activity_prize_item_dict['limit_type'] = limit_type
        activity_prize_item_dict['limit_count'] = limit_count
        activity_prize_item_dict['prize_type'] = prize_type
        activity_prize_item_dict['prize_serial_no'] = prize_serial_no
        activity_prize_item_dict['expiration_date'] = expiration_date
        activity_prize_item_dict['prize_begin_date'] = prize_begin_date
        activity_prize_item_dict['prize_end_date'] = prize_end_date
        activity_prize_item_dict['prize_get_limit_type'] = prize_get_limit_type
        activity_prize_item_dict['prize_get_limit_skus'] = prize_get_limit_skus
        activity_prize_item_list.append(activity_prize_item_dict)
    return activity_prize_item_list

# 写入数据：
def write_prize_data(result_excel_path,prize_data,write_index):
    wb = openpyxl.load_workbook(result_excel_path)
    ws = wb['领券中心标准化检查结果']
    if isinstance(prize_data,dict):
        write_column_index = 1
        if prize_data['cycle_type'] == 'none':
            prize_data['cycle_type'] = '否'
            prize_data['active_dates'] = ''
        else:
            pass
        if prize_data['new_member_able'] == '0':
            prize_data['new_member_able'] = '不限制新客领取'
        else:
            prize_data['new_member_able'] = '仅新客领取'
        if prize_data['visibility'] == 'public':
            prize_data['visibility'] = '商品详情界面可领'
        else:
            prize_data['visibility'] = '仅活动界面可领'
        if prize_data['shareable'] == '1':
            prize_data['shareable'] = '支持分享'
        else:
            prize_data['shareable'] = '不支持分享'
        for data in prize_data.values():
            ws.cell(row= write_index, column= write_column_index, value= data)
            write_column_index += 1
    wb.save(result_excel_path)
    return 0



if __name__ == '__main__':
    org_code = 'jwbaby'
    activity_code = 'gc_1607309295950189'

    excel_name_time = get_time_str()
    # 存放结果的文件的绝对路径
    result_excel_path = r'C:\Users\LyyCc\Desktop\领券中心标准化-%s.xlsx' %excel_name_time
    # 有对应模板的excel文件执行load_excel,如果没有执行create_excel+load_excel
    print("当前工作文件路径为：",result_excel_path)    # 输出当前文件路径
    # 创建文件并写入标准模板
    create_excel(result_excel_path)
    write_medoel_excel(result_excel_path)


    avtivity_base_info_data = query_avtivity_base_info(org_code,activity_code)
    activity_id = avtivity_base_info_data['id']
    avtivity_prize_info_data = query_avtivity_prize_info(org_code,activity_id)
    write_prize_data(result_excel_path,avtivity_base_info_data,write_index='3')

    print("文件生成完毕，请认真检查数据！")
