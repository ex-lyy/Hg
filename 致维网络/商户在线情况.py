# -*- coding:utf-8 -*-
# 文件名称：Lyy-商户在线情况
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/11/2 15:22

from Mymain import *
import openpyxl

# 查询的时间
created_at = '2020-09-01'

if __name__ == '__main__':

    server, dbconfig, cursor = connect_masterDB('ris_production')
    retailer_sql = "SELECT * FROM ris_production.global_retailers WHERE profile_id IN ('1','8','9') AND state = 'online' AND service_provider_code='global' AND code NOT IN ('loveheart','yyplan','beibei','jdb','sunnybaby','harneybaby','jzaibaby','yqmamalove','bzyqyy','xwmaidibaby','fddamuzhi','xcbeierjiazu','xybuyingfang','czaiyingjiayuan','yzbeibeiyy','czhjyy','npfuwawa','jnyft','lxdodoisland','lzsimier','gdxtzgg','aqaxqzy','ssjjw','scayf','aqjbb','xhyyzj','yfxty','szxsbb','yymml','wsygbb','ddjsts','kpdcyy','hdayy');"
    cursor.execute(retailer_sql)
    res = cursor.fetchall()


    org_list = []   # 定义org_code的列表
    name_list = []  # 定义商户名称的列表
    profile_id_list = []    # 定义商户类型的列表

    for res_i in res:
        org_list.append(res_i['code'])
        name_list.append(res_i['name'])
        if res_i['profile_id'] == 1:
            profile_id_list.append('直营')
        elif res_i['profile_id'] == 8:
            profile_id_list.append('联营大户')
        elif res_i['profile_id'] == 9:
            profile_id_list.append('梧桐')


    # 会员不正常的商户：
    unnormal_member_org_list = []

    # 订单奖励数量不正常的商户：
    unnormal_prize_org_list = []

    # 促销活动奖励数量不正常的商户：
    unnormal_promotion_org_list = []

    # 抽奖活动数量不正常的商户：
    unnormal_marketing_org_list = []


    for org_code in org_list:
        # 近三个月未注册会员的信息
        register_state = "SELECT COUNT(1) as member_count FROM ris_production.members WHERE org_code='%s' AND created_at>='%s';"%(org_code,created_at)
        cursor.execute(register_state)
        member_count = cursor.fetchone()
        if member_count['member_count'] >= 3:
            print("%s注册会员数量正常"%org_code)
        else:
            unnormal_member_org_list.append(org_code)

    for org_code in org_list:
        # 近三个月没有订单奖励活动的信息
        prize_activity_state = "SELECT COUNT(1) as prize_count FROM pomelo_backend_production.prize_activities WHERE org_code='%s' AND created_at>='%s';" %(org_code,created_at)
        cursor.execute(prize_activity_state)
        prize_count = cursor.fetchone()
        if prize_count['prize_count'] >= 1:
            print("%s订单奖励信息正常" % org_code)
        else:
            unnormal_prize_org_list.append(org_code)

    for org_code in org_list:
        # 近三个月没有互动活动的信息
        marketing_state = "SELECT COUNT(1) as marketing_count FROM pomelo_backend_production.marketing_prize_events WHERE org_code='%s' AND created_at>='%s';"%(org_code,created_at)
        cursor.execute(marketing_state)
        marketing_count = cursor.fetchone()
        if marketing_count['marketing_count'] >= 1:
            print("%s抽奖活动信息正常" % org_code)
        else:
            unnormal_marketing_org_list.append(org_code)


    for org_code in org_list:
        # 近三个月没有促销活动的信息
        pormotion_state = "SELECT COUNT(1) as pormotion_count FROM pomelo_backend_production.promotion_promotions WHERE org_code='%s' AND created_at>='%s';" %(org_code,created_at)
        cursor.execute(pormotion_state)
        pormotion_count = cursor.fetchone()
        if pormotion_count['pormotion_count'] >= 1:
            print("%s促销活动数量正常" % org_code)
        else:
            unnormal_promotion_org_list.append(org_code)


    unnormal_list_list = [unnormal_member_org_list,unnormal_prize_org_list,unnormal_promotion_org_list,unnormal_marketing_org_list]

    #设置放置异常商户列表，记录长度，方便比较出拥有最大元素数量的列表
    unnormal_length_list = []
    # 获取四个列表长度放在另一个列表里，方便比较什么列表元素最多
    for unnormal_list in unnormal_list_list:
        unnormal_length_list.append(len(unnormal_list))

    max_length = max(unnormal_length_list)  #获取四个列表的元素最大数量

    max_index = unnormal_length_list.index(max_length)  #获取元素数量最多的列表的索引，反向定位到是哪一个列表

    max_unnormal_list = unnormal_list_list.pop(max_index)   #将元素最多的列表拿出来，用于全元素比较

    real_unnormal_org_list = [] #定义真正异常的商户

    # 只有会员注册、订单奖励、促销活动、抽奖活动均异常的商户我们才认为是异常商户
    for  max_unnormal_list_i in max_unnormal_list:
        if (max_unnormal_list_i in unnormal_list_list[0]) & (max_unnormal_list_i in unnormal_list_list[1]) & (max_unnormal_list_i in unnormal_list_list[2]):
            real_unnormal_org_list.append(max_unnormal_list_i)




    close_sshserver(server, dbconfig, cursor)


    Excel_path = r'C:\Users\LyyCc\Desktop\商户在线情况'+created_at+r'.xlsx'

    # 先创建一个表格
    Excel_path = r'C:\Users\LyyCc\Desktop\商户在线情况' + created_at + r'.xlsx'

    wb_create = openpyxl.Workbook(Excel_path)
    ws_sheet1_create = wb_create.create_sheet('Sheet1')
    ws_sheet2_create = wb_create.create_sheet('Sheet2')
    wb_create.save(Excel_path)

    # 往表格写入数据
    wb = openpyxl.load_workbook(Excel_path)
    ws_sheet1 = wb['Sheet1']
    ws_sheet2 = wb['Sheet2']

    # 写表格的头
    ws_sheet1.cell(row = 1,column = 1,value='商户code')
    ws_sheet1.cell(row = 1,column = 2,value='商户名称')
    ws_sheet1.cell(row = 1,column = 3,value='商户类型')
    ws_sheet1.cell(row = 1,column = 4,value='异常商户code')
    ws_sheet1.cell(row = 1,column = 5,value='异常商户名称')
    ws_sheet1.cell(row = 1,column = 6,value='异常商户类型')

    # 写入正常商户code、名称、商户类型
    for index_org in range(2,len(org_list)+2):
        ws_sheet1.cell(row = index_org,column = 1, value = org_list[index_org-2])
        ws_sheet1.cell(row = index_org,column = 2, value = name_list[index_org-2])
        ws_sheet1.cell(row = index_org,column = 3, value = profile_id_list[index_org-2])

    # 写入异常商户code、名称、商户类型
    for real_unnormal_index in range(2,len(real_unnormal_org_list)+2):
        ws_sheet1.cell(row = real_unnormal_index,column = 4, value = real_unnormal_org_list[real_unnormal_index-2])
        # 找到该商户在商户code列表中的索引，去名称列表里面拿到对应的商户名称
        org_name_index = org_list.index(real_unnormal_org_list[real_unnormal_index-2])
        ws_sheet1.cell(row = real_unnormal_index,column = 5, value = name_list[org_name_index])
        ws_sheet1.cell(row = real_unnormal_index,column = 6, value = profile_id_list[org_name_index])

    # 写表格的头
    ws_sheet2.cell(row = 1,column = 1,value='三个月内少于三个会员注册')
    ws_sheet2.cell(row = 1,column = 3,value='三个月内没有上线订单奖励')
    ws_sheet2.cell(row = 1,column = 5,value='三个月内没有上线促销活动')
    ws_sheet2.cell(row = 1,column = 7,value='三个月内没有上线抽奖活动')
    ws_sheet2.cell(row = 1,column = 2,value='三个月内少于三个会员注册-名称')
    ws_sheet2.cell(row = 1,column = 4,value='三个月内没有上线订单奖励-名称')
    ws_sheet2.cell(row = 1,column = 6,value='三个月内没有上线促销活动-名称')
    ws_sheet2.cell(row = 1,column = 8,value='三个月内没有上线抽奖活动-名称')


    # 写入会员注册异常的商户数据
    for unnormal_member_org_index in range(2,len(unnormal_member_org_list)+2):
        ws_sheet2.cell(row=unnormal_member_org_index, column=1, value=unnormal_member_org_list[unnormal_member_org_index-2])
        # 找到该商户在商户code列表中的索引，去名称列表里面拿到对应的商户名称
        unnormal_member_org_name_index = org_list.index(unnormal_member_org_list[unnormal_member_org_index-2])
        ws_sheet2.cell(row=unnormal_member_org_index, column=2, value=name_list[unnormal_member_org_name_index])

    # 写入订单奖励异常的商户数据
    for unnormal_prize_org_index in range(2,len(unnormal_prize_org_list)+2):
        ws_sheet2.cell(row=unnormal_prize_org_index, column=3, value=unnormal_prize_org_list[unnormal_prize_org_index-2])
        # 找到该商户在商户code列表中的索引，去名称列表里面拿到对应的商户名称
        unnormal_prize_org_name_index = org_list.index(unnormal_prize_org_list[unnormal_prize_org_index-2])
        ws_sheet2.cell(row=unnormal_prize_org_index, column=4, value=name_list[unnormal_prize_org_name_index])

    # 写入促销活动异常的商户数据
    for unnormal_promotion_org_index in range(2,len(unnormal_promotion_org_list)+2):
        ws_sheet2.cell(row=unnormal_promotion_org_index, column=5, value=unnormal_promotion_org_list[unnormal_promotion_org_index-2])
        # 找到该商户在商户code列表中的索引，去名称列表里面拿到对应的商户名称
        unnormal_promotion_org_name_index = org_list.index(unnormal_promotion_org_list[unnormal_promotion_org_index-2])
        ws_sheet2.cell(row=unnormal_promotion_org_index, column=6, value=name_list[unnormal_promotion_org_name_index])

    # 写入抽奖活动异常的商户数据
    for unnormal_marketing_org_index in range(2,len(unnormal_marketing_org_list)+2):
        ws_sheet2.cell(row=unnormal_marketing_org_index, column=7, value=unnormal_marketing_org_list[unnormal_marketing_org_index-2])
        # 找到该商户在商户code列表中的索引，去名称列表里面拿到对应的商户名称
        unnormal_marketing_org_name_index = org_list.index(unnormal_marketing_org_list[unnormal_marketing_org_index-2])
        ws_sheet2.cell(row=unnormal_marketing_org_index, column=8, value=name_list[unnormal_marketing_org_name_index])

    wb.save(Excel_path) #保存Excel文件