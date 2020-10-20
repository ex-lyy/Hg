# -*- encoding：utf-8 -*-
# 文件名称：Lyy-小程序表格合并
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/9/23 上午 10:45


import openpyxl


# 获取原始表格数据信息
def get_one_information(one_excel_path):
    # 合并之后的表格路径
    three_excel_path = r"C:\Users\LyyCc\Desktop\商户小程序合并.xlsx"
    wb1 = openpyxl.load_workbook(one_excel_path)
    ws1 = wb1['Sheet1']
    wb2 = openpyxl.load_workbook(three_excel_path)
    ws2 = wb2['Sheet1']
    # 输出表格最大行数用于遍历
    max_row1 = ws1.max_row
    for i in range(1,max_row1):
        if ws1.cell(row=i + 2, column=3).value == '叶云飞':
        for j in range(17):



            print(ws1.cell(row=i + 2, column=j + 1).value, end='、')
        print('')
    return 0


# 获取小程序配置表格数据信息
def get_two_information(two_excel_path):
    wb = openpyxl.load_workbook(two_excel_path)
    ws = wb['Sheet1']
    # 输出表格最大行数用于遍历
    max_row = ws.max_row
    print("表格2最大行数:", max_row)
    return 0


# 写入合并表格数据信息
def write_three_information(three_excel_path):
    wb = openpyxl.load_workbook(three_excel_path)
    ws = wb['Sheet1']
    # 输出表格最大行数用于遍历
    print(ws.max_row)
    return 0


if __name__ == '__main__':
    # 玉婷给的原始表格路径
    one_excel_path = r"C:\Users\LyyCc\Desktop\新版小程序2.1.15版本预推送结果20200725.xlsx"
    # 小程序配置的表格路径
    two_excel_path = r"C:\Users\LyyCc\Desktop\商户小程序信息.xlsx"
    # 合并之后的表格路径
    three_excel_path = r"C:\Users\LyyCc\Desktop\商户小程序合并.xlsx"
    get_one_information(one_excel_path)
    # get_two_information(two_excel_path)
    # write_three_information(three_excel_path)
