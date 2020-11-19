# -*- coding:utf-8 -*-
# 文件名称：Lyy-批量发优惠券
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/11/10 10:51

'''
Readme:
    此程序是将Excel表格里面的会员数据进行处理成发券脚本
    Excel表格里面3456列分别为会员号、券号、开始时间、结束时间
    此项会生成一个txt文件，将脚本在pomelo控制台执行即可
'''


import time
import openpyxl
import random

def write_code(code_path,member_no,serial_no,active_date,expiration_date,rows_index):
    time_time = get_time()+str(rows_index)
    org_code = 'gynwbb'
    source_key = time_time
    code_text = "NormalActivity::Action::AddCoupon.new(org_code: '%s',member_no: '%s',:source_type => 'bufa', :source_key => '%s', :source => 'pomelo',params: [{serial_no: '%s', state: 'active', active_date: Date.parse('%s'), expiration_date: Date.parse('%s')}]).go" % (
    org_code, member_no, source_key, serial_no, active_date, expiration_date)
    with open(code_path,'a+') as code_file:
        code_file.write(code_text)
        code_file.write('\n')

# 获取当前时间戳作为source_key,这里的source_key作为
def get_time():
    rand_time = random.randint(0,100000000)
    time_time = "cc20201110" + str(rand_time)
    return time_time


if __name__ == '__main__':
    Excel_path = r'C:\Users\LyyCc\Desktop\发券的会员数据.xlsx'
    code_path = r'C:\Users\LyyCc\Desktop\gynwbb会员发券脚本.txt'

    code_text_list = []

    wb = openpyxl.load_workbook(Excel_path)
    ws = wb['Sheet1']

    max_rows = ws.max_row
    max_columns = ws.max_column

    for rows_index in range(2, max_rows + 1):
        member_no = ws.cell(row=rows_index, column=3).value
        serial_no = ws.cell(row=rows_index, column=4).value
        # 此项注释掉的代码是应对某个商户的特殊需求，不用理会
        # if serial_no == '030100300008':
        #     serial_no='T20000182'
        # elif serial_no == '030100300009':
        #     serial_no='T20000202'
        # elif serial_no == '030100300010':
        #     serial_no='T20000203'
        # elif serial_no == '030100300011':
        #     serial_no='T20000183'
        # elif serial_no == '030100600001':
        #     serial_no='T20000188'
        # elif serial_no == '030100600004':
        #     serial_no='T20000189'
        # elif serial_no == '030100600019':
        #     serial_no='T20000190'
        # elif serial_no == '030100600030':
        #     serial_no='T20000187'
        # elif serial_no == '030100600061':
        #     serial_no='T20000188'
        # elif serial_no == '030100600065':
        #     serial_no='T20000204'
        # elif serial_no == '030100600076':
        #     serial_no='T20000191'
        # elif serial_no == '030100600077':
        #     serial_no='T20000189'
        active_date = ws.cell(row=rows_index, column=5).value
        expiration_date = ws.cell(row=rows_index, column=6).value
        write_code(code_path,member_no,serial_no,active_date,expiration_date,rows_index)


