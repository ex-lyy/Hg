# -*- encoding：utf-8 -*-
# 文件名称：Lyy-配置任务数量统计
# 作者:清许丶
# 谨记:遵守秩序，尊重选择。
# 创建时间：2020/8/27 15:33
import openpyxl
import re


def write_excel_index_data(path):
    read_wb = openpyxl.load_workbook(path)
    read_ws_first = read_wb.get_sheet_by_name('Sheet1')
    read_ws_second = read_wb.get_sheet_by_name('Sheet2')
    read_ws_third = read_wb.get_sheet_by_name('Sheet3')
    first_index_date_list = ['日期', '上线业务数', '类型', '次数', '商户', '业务上线数', '在线运营直营商户数', '业务上线商户数', '业务上线率', '前五业务',
                             '前五业务占比', '在线运营的梧桐商户数']
    second_index_date_list = ['商户', '业务名称', '类型', '时间', '上线人', '检查人']
    third_index_date_list = ['商户', '任务名称', 'AM', '配置人员', '互检人员', '操作时间']
    check_opposite_list = ['维成、雅馨-周猛互检', '王歆雨、陈冲-张波互检', '骆翔、张波、张嘉媛-陈冲互检', '虬珅、周猛-姝锋互检', '邹庆、宇豪-俊奇互检', '姝锋-小甘互检',
                           '奇成-施互检', '苏航、施-东互检', '辛迟明-宇豪互检']
    for write_flag1 in range(len(first_index_date_list)):
        read_ws_first.cell(row=1, column=write_flag1 + 1, value=first_index_date_list[write_flag1])
    for write_flag2 in range(len(second_index_date_list)):
        read_ws_second.cell(row=1, column=write_flag2 + 1, value=second_index_date_list[write_flag2])
    for write_flag3 in range(len(third_index_date_list)):
        read_ws_third.cell(row=1, column=write_flag3 + 1, value=third_index_date_list[write_flag3])
def write_excel_data(path):
    read_wb = openpyxl.load_workbook(path)
    read_ws_first = read_wb.get_sheet_by_name('Sheet1')

    pass


def deal_txt_file(path):
    dates = []
    org_name_list = []
    operate_name_list = []
    business_type_list = []
    with open(path, 'r', encoding='utf8') as txt_file:
        txt_file_lines = txt_file.readlines()
        for txt_file_line in txt_file_lines:
            print('当前正在操作第%s行的数据。' % (txt_file_lines.index(txt_file_line) + 1))
            if not txt_file_line:
                break
            date = re.match(r'\d{2}.\d{2}', txt_file_line)
            if date:
                dates.append(date.group())
            else:
                org_name = re.match(r'【[\u4e00-\u9fa5]*】', txt_file_line).group()
                operate_name = re.search(r' [\u4e00-\u9fa5]*', txt_file_line).group()
                business_type = re.search(r' x[\u4e00-\u9fa5,\S]*', txt_file_line).group()
                org_name_list.append(org_name.replace(' ', '').replace('【', '').replace('】', ''))
                operate_name_list.append(operate_name.replace(' ', ''))
                business_type_list.append(business_type.replace(' ', '').replace('x', ''))
    if len(org_name_list) == len(operate_name_list) == len(business_type_list):
        print("*" * 30 + "数据量校验正确" + "*" * 30)
    return dates, org_name_list, operate_name_list, business_type_list


if __name__ == '__main__':
    # excel_path = r'C:\Users\LyyCc\Desktop\0817-0823配置统计.xlsx'
    text_path = r'C:\Users\LyyCc\Desktop\配置任务统计.txt'
    write_excel_path_first = r'C:\Users\LyyCc\Desktop'
    dates, org_name_list, operate_name_list, business_type_list = deal_txt_file(text_path)
